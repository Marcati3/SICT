"""Extract all dashboard data from the 3 source spreadsheets for v11 update."""
import os, json, sys, io
from collections import defaultdict
from datetime import datetime

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl

DL = os.path.expanduser('~/Downloads')

# ── 1. Region summary from File 1 ──
wb1 = openpyxl.load_workbook(os.path.join(DL, 'Total Revenue YTD by Region (WR-8).xlsx'), data_only=True)
ws1 = wb1['Export']
regions = {}
for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row, values_only=True):
    name = row[0]
    if not name or name in ('', 'Applied filters:'):
        continue
    regions[name] = {
        'target': row[1] or 0,
        'ytd': row[2] or 0,
        'so': row[3] or 0,
        'ytd_so': row[4] or 0,
        'gap': row[5] or 0,
        'pct_ach': row[6] or 0,
        'vs_last_wk': row[9] or 0,
    }
wb1.close()

print("=== REGION SUMMARY ===")
for k, v in regions.items():
    print(f"  {k}: YTD=${v['ytd']:,.0f}  SO=${v['so']:,.0f}  YTD+SO=${v['ytd_so']:,.0f}  Target=${v['target']:,.0f}  Gap=${v['gap']:,.0f}  %Ach={v['pct_ach']:.1%}")

# ── 2. Q1 summary from File 2 ──
wb2 = openpyxl.load_workbook(os.path.join(DL, 'Quarter Revenue Update (M$) (WR1).xlsx'), data_only=True)
ws2 = wb2['Export']
row2 = list(ws2.iter_rows(min_row=2, max_row=2, values_only=True))[0]
q1 = {
    'target': row2[0] or 0,
    'qtd': row2[1] or 0,
    'so': row2[2] or 0,
    'qtd_so': row2[3] or 0,
    'gap': row2[4] or 0,
    'pct_ach': row2[5] or 0,
    'vs_last_wk': row2[6] or 0,
    'pct_progress': row2[7] or 0,
    'sc_confirm': row2[8] or 0,
}
wb2.close()

print("\n=== Q1 SUMMARY ===")
print(f"  Target: ${q1['target']:,.0f}")
print(f"  QTD Actual: ${q1['qtd']:,.0f}")
print(f"  Open SO: ${q1['so']:,.0f}")
print(f"  QTD+SO: ${q1['qtd_so']:,.0f}")
print(f"  Gap: ${q1['gap']:,.0f}")
print(f"  %Ach: {q1['pct_ach']:.1%}")

# ── 3. Revenue by salesperson/month/customer/product from File 3 Revenue sheet ──
path3 = os.path.join(DL, 'Revenue, SO, Rolling Forecast and Sale Budget Comparison.xlsx')
wb3 = openpyxl.load_workbook(path3, data_only=True)

# Revenue (actuals) - filter to 2026 only
ws_rev = wb3['Revenue']
rev_by_person = defaultdict(float)
rev_by_month = defaultdict(float)
rev_by_customer = defaultdict(float)
rev_by_product = defaultdict(float)
rev_by_region = defaultdict(float)

MONTH_MAP = {
    'January': 'Jan', 'February': 'Feb', 'March': 'Mar', 'April': 'Apr',
    'May': 'May', 'June': 'Jun', 'July': 'Jul', 'August': 'Aug',
    'September': 'Sep', 'October': 'Oct', 'November': 'Nov', 'December': 'Dec'
}

# Salesperson name normalization
def norm_sales(name):
    if not name:
        return 'Unknown'
    n = str(name).strip()
    mapping = {
        'Intira Loychoosak': 'Intira', 'Intira': 'Intira',
        'LI SHA LI': 'Lisa', 'Lisa': 'Lisa', 'LI SHA': 'Lisa',
        'Nuttapon Panyarachun': 'Nuttapon', 'Nuttapon': 'Nuttapon',
        'Shan Jiang': 'Shan', 'Shan': 'Shan',
        'Terry': 'Terry', 'Terry Chen': 'Terry',
        'Tunn Prasoprat': 'Tunn', 'Tunn': 'Tunn',
    }
    for key, val in mapping.items():
        if key.lower() in n.lower():
            return val
    return n

count = 0
for row in ws_rev.iter_rows(min_row=2, max_row=ws_rev.max_row, values_only=True):
    year = row[34]  # Year column
    if year != 2026:
        continue
    month = str(row[1]) if row[1] else ''
    usd = row[16] or 0  # USD Amount
    if usd == 0:
        continue
    person = norm_sales(row[20])  # Sales Nick Name
    customer = str(row[23]) if row[23] else 'Unknown'  # CustShortName
    product = str(row[8]) if row[8] else 'Unknown'  # Product Code
    region = str(row[24]) if row[24] else 'Unknown'  # Region

    short_month = MONTH_MAP.get(month, month[:3] if month else '')

    rev_by_person[person] += usd
    rev_by_month[short_month] += usd
    rev_by_customer[customer] += usd
    rev_by_product[product] += usd
    rev_by_region[region] += usd
    count += 1

print(f"\n=== REVENUE 2026 ({count} rows) ===")
print("\nBy Salesperson:")
for k in sorted(rev_by_person, key=rev_by_person.get, reverse=True):
    print(f"  {k}: ${rev_by_person[k]:,.0f}")
print("\nBy Month:")
for m in ['Jan', 'Feb', 'Mar']:
    print(f"  {m}: ${rev_by_month.get(m, 0):,.0f}")
print("\nTop 10 Customers:")
for k in sorted(rev_by_customer, key=rev_by_customer.get, reverse=True)[:10]:
    print(f"  {k}: ${rev_by_customer[k]:,.0f}")
print("\nTop 10 Products:")
for k in sorted(rev_by_product, key=rev_by_product.get, reverse=True)[:10]:
    print(f"  {k}: ${rev_by_product[k]:,.0f}")
print("\nBy Region:")
for k in sorted(rev_by_region, key=rev_by_region.get, reverse=True):
    print(f"  {k}: ${rev_by_region[k]:,.0f}")

# ── 4. SO pending - filter to outstanding > 0 and 2026 shipment ──
ws_so = wb3['SO pending']
so_by_person = defaultdict(float)
so_by_month = defaultdict(float)
so_by_customer = defaultdict(float)
so_by_product = defaultdict(float)

so_count = 0
for row in ws_so.iter_rows(min_row=2, max_row=ws_so.max_row, values_only=True):
    year = row[18]  # Shipment (Years)
    if year != 2026:
        continue
    outstanding = row[15] or 0  # Outstanding QTY
    if outstanding <= 0:
        continue
    usd = row[47] or 0  # Amount ($)
    if usd == 0:
        # Try SUM USD
        usd = row[25] or 0
    if usd == 0:
        continue

    month = str(row[19]) if row[19] else ''  # Shipment (Month)
    person = norm_sales(row[28])  # Sales Nick Name
    customer = str(row[2]) if row[2] else str(row[1]) if row[1] else 'Unknown'  # Customer Short Name
    product = str(row[7]) if row[7] else 'Unknown'  # Product code

    short_month = MONTH_MAP.get(month, month[:3] if month else '')

    so_by_person[person] += usd
    so_by_month[short_month] += usd
    so_by_customer[customer] += usd
    so_by_product[product] += usd
    so_count += 1

print(f"\n=== OPEN SO 2026 ({so_count} rows) ===")
print("\nBy Salesperson:")
for k in sorted(so_by_person, key=so_by_person.get, reverse=True):
    print(f"  {k}: ${so_by_person[k]:,.0f}")
print("\nBy Month:")
for m in ['Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']:
    v = so_by_month.get(m, 0)
    if v > 0:
        print(f"  {m}: ${v:,.0f}")
print("\nTop 10 Customers (SO):")
for k in sorted(so_by_customer, key=so_by_customer.get, reverse=True)[:10]:
    print(f"  {k}: ${so_by_customer[k]:,.0f}")
print("\nTop 10 Products (SO):")
for k in sorted(so_by_product, key=so_by_product.get, reverse=True)[:10]:
    print(f"  {k}: ${so_by_product[k]:,.0f}")

# ── 5. Rolling Forecast ──
ws_fc = wb3['Rolling Forecast']
fc_by_person = defaultdict(float)
fc_by_month = defaultdict(float)

for row in ws_fc.iter_rows(min_row=2, max_row=ws_fc.max_row, values_only=True):
    if row[2] != 2026:
        continue
    doc_type = str(row[1]) if row[1] else ''
    if doc_type != 'Sales':
        continue
    usd = row[13] or 0  # Amount ($)
    month = str(row[3]) if row[3] else ''
    person = norm_sales(row[6])

    fc_by_person[person] += usd
    fc_by_month[month] += usd

print("\n=== ROLLING FORECAST (Sales type only) ===")
print("\nBy Salesperson:")
for k in sorted(fc_by_person, key=fc_by_person.get, reverse=True):
    print(f"  {k}: ${fc_by_person[k]:,.0f}")
print("\nBy Month:")
for m in ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']:
    v = fc_by_month.get(m, 0)
    if v > 0:
        print(f"  {m}: ${v:,.0f}")

# ── 6. Sales Budget ──
ws_bud = wb3['Sales Budget']
bud_by_person = defaultdict(float)
bud_by_month = defaultdict(float)

for row in ws_bud.iter_rows(min_row=2, max_row=ws_bud.max_row, values_only=True):
    if row[2] != 2026:
        continue
    usd = row[12] or 0  # Amount
    month = str(row[3]) if row[3] else ''
    person = norm_sales(row[6])

    short_month = MONTH_MAP.get(month, month[:3] if month else '')

    bud_by_person[person] += usd
    bud_by_month[short_month] += usd

print("\n=== SALES BUDGET ===")
print("\nBy Salesperson:")
for k in sorted(bud_by_person, key=bud_by_person.get, reverse=True):
    print(f"  {k}: ${bud_by_person[k]:,.0f}")
print("\nBy Month:")
for m in ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']:
    v = bud_by_month.get(m, 0)
    if v > 0:
        print(f"  {m}: ${v:,.0f}")
print(f"\n  TOTAL: ${sum(bud_by_month.values()):,.0f}")

wb3.close()

# ── SUMMARY FOR DASHBOARD ──
total_ytd = regions.get('Total', {}).get('ytd', 0)
total_so = regions.get('Total', {}).get('so', 0)
total_target = regions.get('Total', {}).get('target', 0)

print("\n" + "="*60)
print("DASHBOARD KEY METRICS SUMMARY")
print("="*60)
print(f"Annual Target: ${total_target:,.0f}")
print(f"YTD Revenue: ${total_ytd:,.0f}")
print(f"Total Open SO: ${total_so:,.0f}")
print(f"YTD + SO: ${total_ytd + total_so:,.0f}")
print(f"Gap to Target: ${total_target - total_ytd - total_so:,.0f}")
print(f"Q1 Target: ${q1['target']:,.0f}")
print(f"Q1 QTD: ${q1['qtd']:,.0f}")
print(f"Q1 Open SO: ${q1['so']:,.0f}")
print(f"Q1 QTD+SO: ${q1['qtd_so']:,.0f}")
print(f"Q1 Gap: ${q1['gap']:,.0f}")
print(f"80% FCST total: ${sum(fc_by_person.values()):,.0f}")
print(f"Budget total: ${sum(bud_by_month.values()):,.0f}")
